import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# 掌握修改单元格格式：字体、大小、颜色、填充、数据格式

current_path = os.path.dirname(__file__)
output_path = os.path.join(current_path, 'Tutorial_Output')
save_path = os.path.join(output_path, 'tutorial04.xlsx')
os.makedirs(output_path, exist_ok=True)

wb = Workbook()
ws = wb.active


# font_style = Font(size=16, bold=True)  # italic=True, color="FF0000", name="Arial"
# ws['A1'].font = font_style
alighment_style = Alignment(horizontal='center', vertical='center')  # wrap_text=True, text_rotation=90
ws['A1'].alignment = alighment_style
ws['A1'] = 'PSI'




wb.save(save_path)


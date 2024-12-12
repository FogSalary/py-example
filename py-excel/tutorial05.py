import os
from openpyxl import Workbook

# 掌握合并单元格

current_path = os.path.dirname(__file__)
output_path = os.path.join(current_path, 'Tutorial_Output')
save_path = os.path.join(output_path, 'tutorial05.xlsx')
os.makedirs(output_path, exist_ok=True)


wb = Workbook()
ws = wb.active

ws.merge_cells('A1:D2')
ws.merge_cells('A5:C7')


wb.save(save_path)

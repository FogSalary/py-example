import os
from openpyxl import Workbook


### 掌握选择表单、添加删除表单、修改表单标题、复制表单


current_path = os.path.dirname(__file__)
output_path = os.path.join(current_path, 'Tutorial_Output')
save_path = os.path.join(output_path, 'tutorial02.xlsx')
os.makedirs(output_path, exist_ok=True)


wb = Workbook()

## choose sheet
ws = wb.active
ws['A1'] = 12


## add sheet
print("当前工作簿中含有表单：", wb.sheetnames)
wb.create_sheet("first", 0)  # insert at first position
wb.create_sheet("end")  # insert at the end (default)
wb.create_sheet("end2", -1)  # insert at the penultimate position
print("当前工作簿中含有表单：", wb.sheetnames)
# for sheet in wb:
    # print(sheet.title)
# now you can see the output in terminal, sheetnames is from ['Sheet'] -> ['first', 'Sheet', 'end2', 'end']

## adjust sheet name
ws1 = wb['first']
ws1['A1'] = 15
ws1.title = "new name"


## copy sheet
source_ws = wb['new name']
target_ws = wb.copy_worksheet(source_ws)  # default copy sheet will insert at the end


wb.save(save_path)



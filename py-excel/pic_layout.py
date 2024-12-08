import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image as PILImage


def set_pic_size_in_inch(image_path, width, dpi=96):
    img = Image(image_path)
    # 设置图片宽度（单位：英寸）
    desired_width_in_inches = width # 设置目标宽度为 2 英寸
    # dpi = 96  # 假设每英寸 96 像素（Excel 默认 DPI）
    img.width = desired_width_in_inches * dpi

    # 保持图片宽高比例，调整高度
    with PILImage.open(image_path) as pil_img:
        aspect_ratio = pil_img.height / pil_img.width  # 计算宽高比
    img.height = img.width * aspect_ratio

    return img


current_path = os.path.dirname(__file__)
pic_path = os.path.join(current_path, 'pic.png')
save_path = os.path.join(current_path, 'test_excel2.xlsx')


wb = Workbook()
ws = wb.active
ws.title = "ImageList"


# merge cell
# ws.merge_cells('A1:D2')

# font_style = Font(size=16, bold=True)  # italic=True, color="FF0000", name="Arial"
# ws['A1'].font = font_style
alighment_style = Alignment(horizontal='center', vertical='center')  # wrap_text=True, text_rotation=90
ws['A1'].alignment = alighment_style
ws['A1'] = 'PSI'


# adjust cell size
# ws.column_dimensions['C'].width = 100  # 列宽的单位是字符的宽度（基于默认字体）
# ws.row_dimensions[1].height = 100  # 行高的单位是点（point） 类似字体大小的单位

for col in ['A', 'B', 'C']:
    ws.column_dimensions[col].width = 40


# for row in range(1, 6):
#     ws.row_dimensions[row].height = 40


wb.save(save_path)


# 将列的数字转换为字母
column_number = 31
column_letter = get_column_letter(column_number)
print(f"列号 {column_number} 对应的列字母是：{column_letter}")

# 将列字母转换为数字
column_letter = "E"  # 列字母
column_number = column_index_from_string(column_letter)

print(f"列字母 {column_letter} 对应的列号是：{column_number}")
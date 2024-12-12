import os
from openpyxl import Workbook

# 掌握操作单元格，往单元格中写入数据

current_path = os.path.dirname(__file__)
output_path = os.path.join(current_path, 'Tutorial_Output')
save_path = os.path.join(output_path, 'tutorial03.xlsx')
os.makedirs(output_path, exist_ok=True)


wb = Workbook()
ws = wb.active

# 操作单个单元格
c = ws['A5']
c.value = 16
ws['A4'] = 4
d = ws.cell(row=4, column=2, value=10)

# create 100x100 cells in memory, for nothing
# for x in range(1, 101):
#     for y in range(1, 101):
#         ws.cell(row=x, column=y)

# 操作多个单元格
cell_range = ws['A1':'C2']
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

ws['C9'] = "hello, world"
print(tuple(ws.rows))

print(tuple(ws.columns))


# 只获取单元格的值
for row in ws.values:
    for value in row:
        print(value)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)


# 数据存储 Data storage
c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)



wb.save(save_path)
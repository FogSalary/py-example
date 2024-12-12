import os
from openpyxl import Workbook, load_workbook


current_path = os.path.dirname(__file__)
output_path = os.path.join(current_path, 'Tutorial_Output')
save_path = os.path.join(output_path, 'tutorial01.xlsx')
os.makedirs(output_path, exist_ok=True)

# ============ Part1： 新建工作簿，保存工作簿

# ## 创建工作簿
# wb = Workbook()
# ws = wb.active
# ws['A1'] = 12

# ## 保存工作簿
# wb.save(save_path)


# ============ Part2： 加载现有工作簿，保存工作簿

## 打开工作簿
wb = load_workbook(save_path)
ws = wb.active
ws['A2'] = 13

## 保存工作簿
wb.save(save_path)

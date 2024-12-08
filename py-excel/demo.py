import os
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage


def set_pic_size_in_inch(image_path, width, dpi=96):
    img = Image(image_path)
    # 设置图片宽度（单位：英寸）
    desired_width_in_inches = width # 设置目标宽度为 2 英寸
    # dpi = 96  # 假设每英寸 96 像素（Excel 默认 DPI）
    img.width = desired_width_in_inches * dpi

    # 保持图片宽高比例，调整高度
    with PILImage.open(image_path) as pil_img:
        aspect_ratio = pil_img.height / pil_img.width  # 计算宽高比
    img.height = img.width * aspect_ratio

    return img


current_path = os.path.dirname(__file__)
pic_path = os.path.join(current_path, 'pic.png')
save_path = os.path.join(current_path, 'test_excel.xlsx')

wb = Workbook()
ws = wb.active

print(ws.title)

# create sheet
ws1 = wb.create_sheet("Sheet2")  # insert at the end (default)
ws2 = wb.create_sheet("Sheet3", 0)  # insert at first position
ws3 = wb.create_sheet("Sheet4", -1)  # insert at the penultimate position

ws.title = "New Title"

ws['A1'] = "This is a picture test."

image_path = pic_path
# img = Image(image_path)
# img.width = 100
# img.height = 100
# ws.add_image(img, 'A2')

img = set_pic_size_in_inch(image_path, 3.2)
ws.add_image(img, 'A2')

wb.save(save_path)
